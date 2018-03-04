import openpyxl
import os
import time

while True:
    InputFile = input("Please enter file name (with full path if file is in different folder):")
    #print(os.path.exists(InputFile))
    if os.path.exists(InputFile):
        wb = openpyxl.load_workbook(InputFile)
        break
    print("The file you selected does not exist, please try again")

#print(InputFile)


dict_lookup = {
'á':'a',
'â':'a',
'ậ':'a',
'ä':'a',
'ã':'a',
'ầ':'a',
'à':'a',
'ả':'a',
'ạ':'a',
'ạ':'a',
'å':'a',
'ā':'a',
'ê':'e',
'è':'e',
'ę':'e',
'ế':'e',
'é':'e',
'ế':'e',
'ể':'e',
'ệ':'e',
'ë':'e',
'ě':'e',
'ē':'e',
'é':'e',
'î':'i',
'ï':'i',
'ı':'i',
'ì':'i',
'ĩ':'i',
'ī':'i',
'í':'i',
'ổ':'o',
'ô':'o',
'ợ':'o',
'ồ':'o',
'ø':'o',
'ō':'o',
'ó':'o',
'ö':'o',
'ü':'u',
'ư':'u',
'ù':'u',
'ū':'u',
'ú':'u',
'ứ':'u',
'ç':'c',
'č':'c',
'ş':'c',
'ś':'s',
'š':'s',
'ğ':'g',
'ġ':'g',
'ģ':'g',
'ł':'l',
'ļ':'l',
'ń':'n',
'ñ':'n',
'ň':'n',
'ņ':'n',
'ż':'z',
'ž':'z',
'đ':'d',
'ð':'d',
'ỳ':'y',
'ý':'y',
'ħ':'h',
'ř':'r',
'ŗ':'r',
'ķ':'k',
'æ':'ae',
'ß':'ss',
'þ':'th',
'Á':'A',
'Â':'A',
'Ậ':'A',
'Ä':'A',
'Ã':'A',
'Ầ':'A',
'À':'A',
'Ả':'A',
'Ạ':'A',
'Ạ':'A',
'Å':'A',
'Ā':'A',
'Ê':'E',
'È':'E',
'Ę':'E',
'Ế':'E',
'É':'E',
'Ế':'E',
'Ể':'E',
'Ệ':'E',
'Ë':'E',
'Ě':'E',
'Ē':'E',
'É':'E',
'Î':'I',
'Ï':'I',
'I':'I',
'Ì':'I',
'Ĩ':'I',
'Ī':'I',
'Í':'I',
'Ổ':'O',
'Ô':'O',
'Ợ':'O',
'Ồ':'O',
'Ø':'O',
'Ō':'O',
'Ó':'O',
'Ö':'O',
'Ü':'U',
'Ư':'U',
'Ù':'U',
'Ū':'U',
'Ú':'U',
'Ứ':'U',
'Ç':'C',
'Č':'C',
'Ş':'C',
'Ś':'S',
'Š':'S',
'Ğ':'G',
'Ġ':'G',
'Ģ':'G',
'Ł':'L',
'Ļ':'L',
'Ń':'N',
'Ñ':'N',
'Ň':'N',
'Ņ':'N',
'Ż':'Z',
'Ž':'Z',
'Đ':'D',
'Ð':'D',
'Ỳ':'Y',
'Ý':'Y',
'Ħ':'H',
'Ř':'R',
'Ŗ':'R',
'Ķ':'K',
'Æ':'AE',
'ß':'SS',
'Þ':'TH'
}

def ConvertSpecChars(CellValue, dict_lookup):
    #print(CellValue)
    #newCellVal = CellValue.replace('a','q')
    #print(newCellVal)
    for SpecChar, replacement in dict_lookup.items():
        CellValue = CellValue.replace(SpecChar, replacement)
    return(CellValue)


sheet_list = wb._sheets

for sheet in sheet_list:
    column_list = sheet.columns
    row_list = sheet.rows
    #print(sheet.max_row)
    #print(sheet.max_column)
    for column in range(1,sheet.max_column+1):
        #print(column)
        for row in range(1, sheet.max_row+1):
            #print(row)
            #print(sheet.cell(row=row, column=column).value)
            NewCellValue=ConvertSpecChars(sheet.cell(row=row, column=column).value, dict_lookup)
            #print(NewCellValue)
            sheet.cell(row=row, column=column).value=NewCellValue
wb.save(InputFile)

print("")
print("File was successfully converted!")
time.sleep(5)

